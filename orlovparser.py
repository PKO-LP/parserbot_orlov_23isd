import logging
import time
import requests
import openpyxl
import xlrd
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes

TOKEN = "8624884828:AAGjwQvWM2mtL-GvDsbCfX9ko5cQNyNhfHE"
CACHE_TTL = 600

DEFAULT_PARAMS = {
    "group_row": 3,
    "first_group_col": 3,
    "group_step": 2,
    "first_data_row": 4,
    "day_col": 1,
    "time_col": 2
}

YAROSLAVL_PARAMS = {
    "group_row": 2,
    "first_group_col": 3,
    "group_step": 2,
    "first_data_row": 3,
    "day_col": 1,
    "time_col": 2
}

CORPUS_CONFIG = {
    "shenkurskoe": {
        "name": "Шенкурское отделение",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/69ae843943508.xlsx",
        "params": DEFAULT_PARAMS.copy()
    },
    "bibirevo_1": {
        "name": "Бибирево 1 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/69ae84143ba25.xlsx",
        "params": DEFAULT_PARAMS.copy()
    },
    "bibirevo_2": {
        "name": "Бибирево 2 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/69aebb317264c.xlsx",
        "params": DEFAULT_PARAMS.copy()
    },
    "bibirevo_3": {
        "name": "Бибирево 3 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/69ae84268cd3e.xlsx",
        "params": DEFAULT_PARAMS.copy()
    },
    "bibirevo_45": {
        "name": "Бибирево 4-5 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/69ae842ec58ad.xlsx",
        "params": DEFAULT_PARAMS.copy()
    },
    "yaroslavl_1": {
        "name": "Ярославский 1 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/6998081244da6.xlsx",
        "params": YAROSLAVL_PARAMS.copy()
    },
    "yaroslavl_2": {
        "name": "Ярославский 2 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/6998081e7e260.xlsx",
        "params": YAROSLAVL_PARAMS.copy()
    },
    "yaroslavl_3": {
        "name": "Ярославский 3 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/6998082a24c1f.xlsx",
        "params": YAROSLAVL_PARAMS.copy()
    },
    "yaroslavl_4": {
        "name": "Ярославский 4 курс",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/6998083779866.xlsx",
        "params": YAROSLAVL_PARAMS.copy()
    },
    "volokolamsk_svoboda": {
        "name": "Волоколамский и Свобода",
        "url": "https://spo-13.mskobr.ru/attach_files/upload_users_files/6970c938ddc5f.xls",
        "params": {
            "group_row": 8,
            "first_group_col": 3,
            "group_step": 2,
            "first_data_row": 9,
            "day_col": 1,
            "time_col": 2
        }
    }
}

WEEKDAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.FileHandler("bot.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

cache = {}


class XlsSheetWrapper:
    def __init__(self, sheet):
        self.sheet = sheet
        self.max_row = sheet.nrows

    def cell(self, row, column):
        class Cell:
            def __init__(self, value):
                self.value = value

        try:
            value = self.sheet.cell_value(row - 1, column - 1)
            return Cell(value)
        except:
            return Cell(None)


def load_sheet(url, force=False):
    global cache
    now = time.time()
    if not force and url in cache:
        wb, ts = cache[url]
        if now - ts < CACHE_TTL:
            logger.info(f"⚡ Использую кеш для {url}")
            return wb
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        if url.endswith('.xls') and not url.endswith('.xlsx'):
            book = xlrd.open_workbook(file_contents=response.content)
            sheet = book.sheet_by_index(0)
            wrapper = XlsSheetWrapper(sheet)
            cache[url] = (wrapper, now)
            logger.info(f"✅ XLS файл {url} загружен")
            return wrapper
        else:
            wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
            sheet = wb.active
            cache[url] = (sheet, now)
            logger.info(f"✅ XLSX файл {url} загружен")
            return sheet
    except Exception as e:
        logger.error(f"❌ Ошибка загрузки {url}: {e}")
        return None


def get_groups(sheet, params):
    groups = []
    col = params["first_group_col"]
    while col < 100:
        cell = sheet.cell(row=params["group_row"], column=col).value
        if not cell:
            col += params["group_step"]
            continue
        group_name = str(cell).strip()
        if group_name and group_name not in ["Предмет", "Каб", "Каб.", "День", "#"]:
            groups.append(group_name)
        col += params["group_step"]
    return groups


def normalize_day(day_str):
    if not day_str:
        return None
    raw = str(day_str).strip().lower().replace('.', '').replace(' ', '').replace('-', '')
    days_map = {
        "понедельник": ["пн", "пон", "пнд", "1"],
        "вторник": ["вт", "втр", "2"],
        "среда": ["ср", "сред", "3"],
        "четверг": ["чт", "чет", "4"],
        "пятница": ["пт", "пят", "5"],
        "суббота": ["сб", "суб", "6"],
        "воскресенье": ["вс", "воск", "7"]
    }
    for day, variants in days_map.items():
        if raw == day or any(v in raw for v in variants):
            return day.capitalize()
    return None


def get_schedule(sheet, group_name, target_day, params):
    group_col = None
    cabinet_col = None
    col = params["first_group_col"]
    row_for_search = params["group_row"]

    while col < 100:
        cell = sheet.cell(row=row_for_search, column=col).value
        if not cell:
            col += params["group_step"]
            continue
        cell_str = str(cell).strip()
        if cell_str == group_name:
            group_col = col
            cabinet_col = col + 1
            logger.info(f"✅ Найдена группа '{group_name}' в колонке {group_col}")
            break
        col += params["group_step"]

    if group_col is None:
        return f"❌ *Группа '{group_name}' не найдена.*\n\nПопробуйте выбрать другую группу."

    lessons = []
    current_day = None
    found_target_day = False
    day_finished = False

    for row in range(params["first_data_row"], sheet.max_row + 1):
        if day_finished:
            break

        day_cell = sheet.cell(row=row, column=params["day_col"]).value
        if day_cell:
            normalized = normalize_day(day_cell)
            if normalized:
                if found_target_day and normalized != target_day:
                    day_finished = True
                    continue
                current_day = normalized
                if current_day == target_day:
                    found_target_day = True

        if current_day == target_day:
            time_val = sheet.cell(row=row, column=params["time_col"]).value
            subject = sheet.cell(row=row, column=group_col).value
            cabinet = sheet.cell(row=row, column=cabinet_col).value

            if subject and str(subject).strip() and str(subject).strip() not in ["0", "None", "nan"]:
                if len(lessons) > 0 and str(time_val).strip() == "1" and lessons[-1].startswith("⏰ 1"):
                    continue
                lesson = []
                if time_val and str(time_val).strip() and str(time_val).strip() not in ["0", "None", "nan"]:
                    lesson.append(f"⏰ *{time_val}*")
                lesson.append(f"📚 {subject}")
                if cabinet and str(cabinet).strip() and str(cabinet).strip() not in ["0", "None", "nan"]:
                    lesson.append(f"🏫 *Каб. {cabinet}*")
                lessons.append("\n".join(lesson))

    if not lessons:
        return f"📅 *{target_day}*\n\n🏖️ *В этот день занятий нет.*"

    result = f"📅 *{target_day}*\n\n"
    for i, lesson in enumerate(lessons, 1):
        result += f"*{i}.* {lesson}\n\n"

    return result

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    welcome_text = (
        "🎓 *Добро пожаловать в бот расписания!*\n\n"
        "Здесь вы можете узнать расписание занятий для всех корпусов.\n"
        "Выберите корпус:"
    )
    keyboard = [
        [InlineKeyboardButton("Шенкурское отделение", callback_data="corpus_shenkurskoe")],
        [InlineKeyboardButton("Бибирево", callback_data="corpus_bibirevo")],
        [InlineKeyboardButton("Ярославский", callback_data="corpus_yaroslavl")],
        [InlineKeyboardButton("Волоколамский и Свобода", callback_data="corpus_volokolamsk_svoboda")]
    ]
    await update.message.reply_text(
        welcome_text,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def corpus_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    corpus = query.data.replace("corpus_", "")

    if corpus == "shenkurskoe":
        config_key = "shenkurskoe"
        context.user_data['current_config'] = config_key
        url = CORPUS_CONFIG[config_key]["url"]
        sheet = load_sheet(url)
        if not sheet:
            await query.edit_message_text("❌ *Не удалось загрузить расписание.*\nПопробуйте позже.", parse_mode="Markdown")
            return
        groups = get_groups(sheet, CORPUS_CONFIG[config_key]["params"])
        if not groups:
            await query.edit_message_text("❌ *Не найдены группы.*", parse_mode="Markdown")
            return
        keyboard = [[InlineKeyboardButton(g, callback_data=f"group_{g}")] for g in groups]
        await query.edit_message_text(
            f"🎓 *{CORPUS_CONFIG[config_key]['name']}*\n\nВыберите группу:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif corpus == "bibirevo":
        keyboard = [
            [InlineKeyboardButton("1 курс", callback_data="course_bibirevo_1")],
            [InlineKeyboardButton("2 курс", callback_data="course_bibirevo_2")],
            [InlineKeyboardButton("3 курс", callback_data="course_bibirevo_3")],
            [InlineKeyboardButton("4-5 курс", callback_data="course_bibirevo_45")]
        ]
        await query.edit_message_text(
            "🏢 *Бибирево*\n\nВыберите курс:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif corpus == "yaroslavl":
        keyboard = [
            [InlineKeyboardButton("1 курс", callback_data="course_yaroslavl_1")],
            [InlineKeyboardButton("2 курс", callback_data="course_yaroslavl_2")],
            [InlineKeyboardButton("3 курс", callback_data="course_yaroslavl_3")],
            [InlineKeyboardButton("4 курс", callback_data="course_yaroslavl_4")]
        ]
        await query.edit_message_text(
            "🏛 *Ярославский*\n\nВыберите курс:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif corpus == "volokolamsk_svoboda":
        config_key = "volokolamsk_svoboda"
        context.user_data['current_config'] = config_key
        url = CORPUS_CONFIG[config_key]["url"]
        sheet = load_sheet(url)
        if not sheet:
            await query.edit_message_text("❌ *Не удалось загрузить расписание.*\nПопробуйте позже.", parse_mode="Markdown")
            return
        groups = get_groups(sheet, CORPUS_CONFIG[config_key]["params"])
        if not groups:
            await query.edit_message_text("❌ *Не найдены группы.*", parse_mode="Markdown")
            return
        keyboard = [[InlineKeyboardButton(g, callback_data=f"group_{g}")] for g in groups]
        await query.edit_message_text(
            f"🏭 *{CORPUS_CONFIG[config_key]['name']}*\n\nВыберите группу:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def course_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    course_key = query.data.replace("course_", "")
    context.user_data['current_config'] = course_key
    url = CORPUS_CONFIG[course_key]["url"]
    sheet = load_sheet(url)
    if not sheet:
        await query.edit_message_text("❌ *Не удалось загрузить расписание.*\nПопробуйте позже.", parse_mode="Markdown")
        return
    groups = get_groups(sheet, CORPUS_CONFIG[course_key]["params"])
    if not groups:
        await query.edit_message_text("❌ *Не найдены группы.*", parse_mode="Markdown")
        return
    keyboard = [[InlineKeyboardButton(g, callback_data=f"group_{g}")] for g in groups]
    await query.edit_message_text(
        f"🎓 *{CORPUS_CONFIG[course_key]['name']}*\n\nВыберите группу:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def group_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    group = query.data.replace("group_", "")
    context.user_data['selected_group'] = group

    # Создаем кнопки с полными названиями дней, по 2 в ряд
    day_buttons = []
    row = []
    day_emojis = ["", "", "", "", "", "", ""]
    for i, day in enumerate(WEEKDAYS):
        row.append(InlineKeyboardButton(f"{day_emojis[i]} {day}", callback_data=f"day_{day}"))
        if len(row) == 2:
            day_buttons.append(row)
            row = []
    if row:
        day_buttons.append(row)

    await query.edit_message_text(
        f"🎓 *Группа: {group}*\n\n🗓 Выберите день недели:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(day_buttons)
    )


async def day_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    day = query.data.replace("day_", "")
    group = context.user_data.get('selected_group')
    config_key = context.user_data.get('current_config')
    if not group or not config_key:
        await query.edit_message_text("❌ *Сначала выберите группу через меню.*", parse_mode="Markdown")
        return

    await query.edit_message_text("⏳ *Загружаю расписание...*", parse_mode="Markdown")

    url = CORPUS_CONFIG[config_key]["url"]
    sheet = load_sheet(url)
    if not sheet:
        await query.edit_message_text("❌ *Ошибка загрузки расписания.*\nПопробуйте позже.", parse_mode="Markdown")
        return

    schedule = get_schedule(sheet, group, day, CORPUS_CONFIG[config_key]["params"])

    nav_buttons = [
        [
            InlineKeyboardButton("◀️ К дням", callback_data="back_to_days"),
            InlineKeyboardButton("🔄 Другая группа", callback_data="back_to_groups")
        ],
        [InlineKeyboardButton("🔄 Обновить", callback_data="refresh")]
    ]

    await query.edit_message_text(
        schedule,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(nav_buttons)
    )


async def back_to_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    group = context.user_data.get('selected_group')
    if not group:
        await query.edit_message_text("❌ *Ошибка: группа не выбрана.*", parse_mode="Markdown")
        return

    # Создаем кнопки с полными названиями дней, по 2 в ряд
    day_buttons = []
    row = []
    day_emojis = ["", "", "", "", "", "", ""]
    for i, day in enumerate(WEEKDAYS):
        row.append(InlineKeyboardButton(f"{day_emojis[i]} {day}", callback_data=f"day_{day}"))
        if len(row) == 2:
            day_buttons.append(row)
            row = []
    if row:
        day_buttons.append(row)

    await query.edit_message_text(
        f"🎓 *Группа: {group}*\n\n🗓 Выберите день недели:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(day_buttons)
    )


async def back_to_groups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    config_key = context.user_data.get('current_config')
    if not config_key:
        await query.edit_message_text("❌ *Ошибка: конфигурация не выбрана.*\nНачните с /start", parse_mode="Markdown")
        return

    url = CORPUS_CONFIG[config_key]["url"]
    sheet = load_sheet(url)
    if not sheet:
        await query.edit_message_text("❌ *Не удалось загрузить группы.*", parse_mode="Markdown")
        return
    groups = get_groups(sheet, CORPUS_CONFIG[config_key]["params"])
    if not groups:
        await query.edit_message_text("❌ *Не найдены группы.*", parse_mode="Markdown")
        return
    keyboard = [[InlineKeyboardButton(g, callback_data=f"group_{g}")] for g in groups]
    await query.edit_message_text(
        f"🎓 *{CORPUS_CONFIG[config_key]['name']}*\n\nВыберите группу:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def refresh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    config_key = context.user_data.get('current_config')
    if not config_key:
        await query.edit_message_text("❌ *Ошибка: конфигурация не выбрана.*", parse_mode="Markdown")
        return

    await query.edit_message_text("⏳ *Обновляю данные...*", parse_mode="Markdown")

    url = CORPUS_CONFIG[config_key]["url"]
    sheet = load_sheet(url, force=True)
    if not sheet:
        await query.edit_message_text("❌ *Не удалось обновить расписание.*\nПопробуйте позже.", parse_mode="Markdown")
        return

    group = context.user_data.get('selected_group')
    if group:
        # Создаем кнопки с полными названиями дней, по 2 в ряд
        day_buttons = []
        row = []
        day_emojis = ["", "", "", "", "", "", ""]
        for i, day in enumerate(WEEKDAYS):
            row.append(InlineKeyboardButton(f"{day_emojis[i]} {day}", callback_data=f"day_{day}"))
            if len(row) == 2:
                day_buttons.append(row)
                row = []
        if row:
            day_buttons.append(row)
        await query.edit_message_text(
            f"✅ *Данные обновлены!*\n\n🎓 *Группа: {group}*\n\n🗓 Выберите день:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(day_buttons)
        )
    else:
        groups = get_groups(sheet, CORPUS_CONFIG[config_key]["params"])
        if not groups:
            await query.edit_message_text("❌ *Не найдены группы.*", parse_mode="Markdown")
            return
        keyboard = [[InlineKeyboardButton(g, callback_data=f"group_{g}")] for g in groups]
        await query.edit_message_text(
            f"✅ *Данные обновлены!*\n\n🎓 *{CORPUS_CONFIG[config_key]['name']}*\n\nВыберите группу:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


def main():
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(corpus_selected, pattern="^corpus_"))
    app.add_handler(CallbackQueryHandler(course_selected, pattern="^course_"))
    app.add_handler(CallbackQueryHandler(group_selected, pattern="^group_"))
    app.add_handler(CallbackQueryHandler(day_selected, pattern="^day_"))
    app.add_handler(CallbackQueryHandler(back_to_days, pattern="^back_to_days$"))
    app.add_handler(CallbackQueryHandler(back_to_groups, pattern="^back_to_groups$"))
    app.add_handler(CallbackQueryHandler(refresh, pattern="^refresh$"))

    print("╔════════════════════════════════════╗")
    print("║     🚀 Бот расписания запущен!    ║")
    print("╚════════════════════════════════════╝")
    app.run_polling()


if __name__ == "__main__":
    main()